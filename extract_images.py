from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

# Load workbook and the '40БТ' sheet
wb = load_workbook("40BT.xlsx")
sheet = wb["40БТ"]
image_loader = SheetImageLoader(sheet)

for row in range(1, sheet.max_row + 1):
    cell = f"B{row}"
    if image_loader.image_in(cell):
        image = image_loader.get(cell)
        number = sheet[f"A{row}"].value
        if number:
            # Convert to RGB to ensure JPEG compatibility
            image.convert("RGB").save(f"{number}.jpeg", "JPEG")
            print(f"Saved image for question {number}")